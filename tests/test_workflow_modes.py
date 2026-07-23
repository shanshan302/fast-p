from pathlib import Path
import tempfile
import threading
import unittest
from unittest.mock import patch

import openpyxl

from fast_p.models import Settings
from fast_p.workflow import JobRunner


def collection_payload(request):
    return {"result": {"platforms": [{
        "platformId": request.platform,
        "success": True,
        "noData": False,
        "data": [{
            "part_number": request.model,
            "manufacturer": request.brand,
            "product_url": f"https://example.test/{request.model}",
            "price_tiers": [{"quantity": 100, "unit_price": 2.0}],
        }],
    }]}}


class FakeCollector:
    def __init__(self, settings):
        self.calls = []

    def __enter__(self):
        return self

    def __exit__(self, *args):
        pass

    def collect(self, request):
        self.calls.append(request)
        return collection_payload(request)


class FakeScreenshotter:
    calls = []

    def __init__(self, chrome, profile):
        pass

    def __enter__(self):
        return self

    def __exit__(self, *args):
        pass

    def capture(self, request, output):
        self.calls.append(request)
        output.mkdir(parents=True, exist_ok=True)
        target = output / f"{request.filename}.png"
        target.write_bytes(b"png")
        return target.name


class WorkflowModesTest(unittest.TestCase):
    def setUp(self):
        FakeScreenshotter.calls = []
        self.settings = Settings("", "", "/unused/chrome", "/unused/profile")

    @staticmethod
    def input_excel(path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["敦煌SKU", "型号", "标准厂牌", "供货价", "最小起订量"])
        sheet.append(["SKU-1", "ABC123", "ACME", 1.0, 100])
        workbook.save(path)
        workbook.close()

    @staticmethod
    def capture_excel(path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["敦煌SKU", "型号", "商品链接", "截图文件名"])
        sheet.append(["SKU-1", "ABC123", "https://example.test/ABC123", ""])
        workbook.save(path)
        workbook.close()

    def runner(self):
        return JobRunner(self.settings, lambda event: None, threading.Event())

    def test_collect_mode_does_not_start_screenshotter(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            excel = root / "input.xlsx"
            self.input_excel(excel)
            with patch("fast_p.workflow.CollectorWorker", FakeCollector), patch(
                "fast_p.screenshot.Screenshotter", side_effect=AssertionError("不应截图")
            ):
                workbook, _ = self.runner().run(excel, root / "out", ["hqchip"], "collect")
            exported = openpyxl.load_workbook(workbook, read_only=True)
            try:
                self.assertEqual(
                    "https://example.test/ABC123", exported.active.cell(2, 8).value
                )
            finally:
                exported.close()

    def test_screenshot_mode_needs_only_a_url_excel(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            excel = root / "collected.xlsx"
            self.capture_excel(excel)
            with patch("fast_p.screenshot.Screenshotter", FakeScreenshotter):
                workbook, _ = self.runner().run(excel, root / "out", [], "screenshot")
            exported = openpyxl.load_workbook(workbook, read_only=True)
            try:
                sheet = exported.active
                headers = [cell.value for cell in sheet[1]]
                self.assertEqual(1, headers.count("商品链接"))
                self.assertEqual(1, headers.count("截图文件名"))
                self.assertEqual(
                    "SKU-1.png", sheet.cell(2, headers.index("截图文件名") + 1).value
                )
            finally:
                exported.close()
            self.assertTrue((root / "out" / "screenshots" / "SKU-1.png").is_file())

    def test_all_mode_composes_collection_and_screenshot(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            excel = root / "input.xlsx"
            self.input_excel(excel)
            with patch("fast_p.workflow.CollectorWorker", FakeCollector), patch(
                "fast_p.screenshot.Screenshotter", FakeScreenshotter,
            ):
                workbook, _ = self.runner().run(excel, root / "out", ["hqchip"], "all")
            self.assertTrue(workbook.is_file())
            self.assertEqual(1, len(FakeScreenshotter.calls))


if __name__ == "__main__":
    unittest.main()
