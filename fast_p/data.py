import datetime
import hashlib
import json
from pathlib import Path
import sqlite3
from urllib.parse import urlsplit
import zipfile

import openpyxl

from .models import InputRow, ItemResult
from .rules import number


def _column(headers, hints):
    for index, header in enumerate(headers):
        text = str(header or "").lower()
        if any(hint.lower() in text for hint in hints):
            return index
    return None


def load_rows(excel: Path):
    workbook = openpyxl.load_workbook(excel, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    columns = {
        "model": _column(headers, ["型号", "model", "pn", "part"]),
        "sku": _column(headers, ["敦煌sku", "sku", "敦煌"]),
        "brand": _column(headers, ["标准厂牌", "厂牌", "brand", "制造商"]),
        "price": _column(headers, ["供货价", "supply", "单价", "price"]),
        "moq": _column(headers, ["最小起订量", "起订", "moq", "数量"]),
    }
    missing = [name for name in ("model", "brand", "price", "moq") if columns[name] is None]
    if missing:
        workbook.close()
        raise ValueError(f"Excel 缺少必要列：{', '.join(missing)}")

    rows = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        model = str(cells[columns["model"]] or "").strip()
        if not model:
            continue
        sku_index = columns["sku"]
        moq = number(cells[columns["moq"]])
        rows.append(InputRow(
            row_number=row_number,
            sku=str(cells[sku_index] or "").strip() if sku_index is not None else "",
            model=model,
            brand=str(cells[columns["brand"]] or "").strip(),
            supply_price=number(cells[columns["price"]]),
            moq=int(moq) if moq is not None else None,
        ))
    workbook.close()
    return rows


def load_capture_results(excel: Path):
    """读取任何包含商品链接的 Excel；截图能力不依赖采集平台或 fast-cli。"""
    workbook = openpyxl.load_workbook(excel, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    columns = {
        "url": _column(headers, ["商品链接", "product_url", "url", "链接"]),
        "sku": _column(headers, ["敦煌sku", "sku", "敦煌"]),
        "model": _column(headers, ["型号", "model", "pn", "part"]),
        "matched_model": _column(headers, ["匹配型号"]),
        "platform_name": _column(headers, ["匹配平台"]),
        "reason": _column(headers, ["原因"]),
        "screenshot": _column(headers, ["截图文件名"]),
    }
    if columns["url"] is None:
        workbook.close()
        raise ValueError("截图 Excel 缺少“商品链接”列")

    results = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        url = str(cells[columns["url"]] or "").strip()
        parsed = urlsplit(url)
        if parsed.scheme not in {"http", "https"} or not parsed.netloc:
            continue

        def value(name):
            index = columns[name]
            return str(cells[index] or "").strip() if index is not None else ""

        model = value("model") or value("matched_model") or f"row-{row_number}"
        results.append(ItemResult(
            row_number=row_number,
            sku=value("sku"),
            model=model,
            status="OK",
            reason=value("reason") or "等待截图",
            platform_name=value("platform_name"),
            url=url,
            matched_model=value("matched_model") or model,
            screenshot=value("screenshot"),
        ))
    workbook.close()
    if not results:
        raise ValueError("截图 Excel 没有有效的 http/https 商品链接")
    return results


class Store:
    def __init__(self, path: Path, fingerprint: str):
        self.connection = sqlite3.connect(path)
        self.connection.execute("CREATE TABLE IF NOT EXISTS meta (key TEXT PRIMARY KEY, value TEXT NOT NULL)")
        self.connection.execute(
            "CREATE TABLE IF NOT EXISTS items (row_number INTEGER PRIMARY KEY, result_json TEXT NOT NULL)"
        )
        old = self.connection.execute("SELECT value FROM meta WHERE key='fingerprint'").fetchone()
        if old and old[0] != fingerprint:
            self.connection.execute("DELETE FROM items")
        self.connection.execute(
            "INSERT OR REPLACE INTO meta(key, value) VALUES('fingerprint', ?)", (fingerprint,)
        )
        self.connection.commit()

    def get(self, row_number: int):
        row = self.connection.execute(
            "SELECT result_json FROM items WHERE row_number=?", (row_number,)
        ).fetchone()
        return ItemResult.from_dict(json.loads(row[0])) if row else None

    def save(self, result: ItemResult):
        self.connection.execute(
            "INSERT OR REPLACE INTO items(row_number, result_json) VALUES(?, ?)",
            (result.row_number, json.dumps(result.to_dict(), ensure_ascii=False)),
        )
        self.connection.commit()

    def all(self):
        rows = self.connection.execute("SELECT result_json FROM items ORDER BY row_number").fetchall()
        return [ItemResult.from_dict(json.loads(row[0])) for row in rows]

    def close(self):
        self.connection.close()


def fingerprint(excel: Path, platforms: list[str]):
    stat = excel.stat()
    raw = json.dumps([str(excel.resolve()), stat.st_size, stat.st_mtime_ns, platforms])
    return hashlib.sha256(raw.encode()).hexdigest()


def export_results(excel: Path, output: Path, results: list[ItemResult]):
    workbook = openpyxl.load_workbook(excel)
    sheet = workbook.active
    headers = ["匹配型号", "匹配平台", "商品链接", "状态", "原因", "截图文件名", "匹配数量", "匹配价格"]
    existing = {str(cell.value or "").strip(): cell.column for cell in sheet[1]}
    if "商品链接" in existing and "截图文件名" in existing:
        columns = {
            header: existing.get(header, sheet.max_column + offset + 1)
            for offset, header in enumerate(headers)
        }
    else:
        start = sheet.max_column + 1
        columns = {header: start + offset for offset, header in enumerate(headers)}
    for header, column in columns.items():
        sheet.cell(1, column, header)
    for result in results:
        status = "已找到" if result.status == "OK" and not result.screenshot_error else (
            "截图失败" if result.status == "OK" else "未找到"
        )
        values = [
            result.matched_model, result.platform_name, result.url, status,
            result.screenshot_error or result.reason, result.screenshot,
            result.quantity, result.price,
        ]
        for header, value in zip(headers, values):
            sheet.cell(result.row_number, columns[header], value)
    stem = excel.stem if excel.stem.endswith("_比价结果") else f"{excel.stem}_比价结果"
    target = output / f"{stem}.xlsx"
    workbook.save(target)
    return target


def make_export(output: Path, workbook: Path, results: list[ItemResult]):
    report = output / "运行报告.txt"
    counts = {}
    for result in results:
        counts[result.status] = counts.get(result.status, 0) + 1
    report.write_text(
        "\n".join([
            f"完成时间：{datetime.datetime.now():%Y-%m-%d %H:%M:%S}",
            f"总数：{len(results)}",
            *(f"{key}：{value}" for key, value in sorted(counts.items())),
        ]) + "\n",
        encoding="utf-8",
    )
    archive = output / f"{workbook.stem}_完整材料.zip"
    with zipfile.ZipFile(archive, "w", zipfile.ZIP_DEFLATED) as zipped:
        zipped.write(workbook, workbook.name)
        zipped.write(report, report.name)
        screenshot_dir = output / "screenshots"
        if screenshot_dir.exists():
            filenames = {
                result.screenshot for result in results
                if result.status == "OK" and not result.screenshot_error and result.screenshot
            }
            for filename in sorted(filenames):
                screenshot = screenshot_dir / Path(filename).name
                if screenshot.is_file():
                    zipped.write(screenshot, f"screenshots/{screenshot.name}")
    return archive
